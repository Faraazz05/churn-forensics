"""
cohort_analysis.py
==================
Customer Health Forensics System — Phase 3
Cohort pivot tables and Excel export.
"""

import numpy as np
import pandas as pd
from pathlib import Path
from segmentation_engine import add_segment_dimensions, MIN_SEGMENT_SIZE


def build_cohort_table(df, row_dim, col_dim, metric="churn_rate", min_size=MIN_SEGMENT_SIZE):
    df = add_segment_dimensions(df)
    rows = []
    for r in sorted(df[row_dim].dropna().unique()):
        for c in sorted(df[col_dim].dropna().unique()):
            seg = df[(df[row_dim]==r)&(df[col_dim]==c)]
            if len(seg) < min_size: continue
            if metric == "churn_rate":
                val = float(seg["churned"].mean()) if "churned" in seg.columns else None
            elif metric == "revenue_at_risk":
                nc = int(seg["churned"].sum()) if "churned" in seg.columns else 0
                val = round(nc * float(seg["monthly_spend"].mean() if "monthly_spend" in seg.columns else 0) * 12, 0)
            elif metric == "segment_size":
                val = len(seg)
            else:
                val = None
            rows.append({row_dim: r, col_dim: c, metric: val})
    if not rows: return pd.DataFrame()
    return pd.DataFrame(rows).pivot(index=row_dim, columns=col_dim, values=metric).round(4)


def build_trend_cohort(df_snapshots, dimension, metric="churn_rate"):
    df_snapshots = add_segment_dimensions(df_snapshots)
    rows = []
    for month in sorted(df_snapshots["snapshot_month"].unique()):
        df_m = df_snapshots[df_snapshots["snapshot_month"]==month]
        for val in df_m[dimension].dropna().unique():
            seg = df_m[df_m[dimension]==val]
            if len(seg) < MIN_SEGMENT_SIZE: continue
            v = float(seg["churned"].mean()) if metric=="churn_rate" else len(seg)
            rows.append({dimension: val, "month": month, metric: v})
    if not rows: return pd.DataFrame()
    return pd.DataFrame(rows).pivot(index=dimension, columns="month", values=metric).round(4)


def build_all_cohort_tables(df) -> dict:
    df = add_segment_dimensions(df)
    tables = {}
    pairs  = [
        ("plan_type","region","churn_rate"),
        ("plan_type","contract_type","churn_rate"),
        ("behavior_tier","nps_band_label","churn_rate"),
        ("plan_type","region","revenue_at_risk"),
        ("plan_type","contract_type","segment_size"),
    ]
    for r,c,m in pairs:
        if r in df.columns and c in df.columns:
            t = build_cohort_table(df, r, c, m)
            if not t.empty:
                tables[f"{r}_x_{c}__{m}"] = t
    return tables


def export_to_excel(tables: dict, segment_results: list, output_path: Path):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        HEADER_FILL    = PatternFill("solid", fgColor="1D3557")
        HEADER_FONT    = Font(color="FFFFFF", bold=True)
        DEGRADE_FILL   = PatternFill("solid", fgColor="FFE0E0")
        IMPROVE_FILL   = PatternFill("solid", fgColor="E0FFE0")
        ALT_FILL       = PatternFill("solid", fgColor="F5F5F5")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Segment summary sheet
        ws = wb.create_sheet("Segment Summary")
        headers = ["segment_id","churn_rate","previous_churn_rate","churn_delta",
                   "health_status","risk_level","revenue_at_risk","acceleration","exceeds_benchmark"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for i, seg in enumerate(segment_results, 2):
            ws.append([seg.get(h,"") for h in headers])
            status = seg.get("health_status","")
            fill   = DEGRADE_FILL if status=="degrading" else IMPROVE_FILL if status=="improving" else (ALT_FILL if i%2==0 else None)
            if fill:
                for col in range(1, len(headers)+1):
                    ws.cell(row=i, column=col).fill = fill
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col)+4, 30)

        # Cohort table sheets
        for name, pivot in tables.items():
            ws2 = wb.create_sheet(name[:31])
            ws2.append([""] + list(pivot.columns))
            for cell in ws2[1]:
                cell.fill = HEADER_FILL; cell.font = HEADER_FONT
            for i, (idx, row) in enumerate(pivot.iterrows(), 2):
                ws2.append([str(idx)] + [round(v,4) if isinstance(v,float) else v for v in row])
                if i%2==0:
                    for col in range(1, len(pivot.columns)+2):
                        ws2.cell(row=i, column=col).fill = ALT_FILL
            for col in ws2.columns:
                ws2.column_dimensions[col[0].column_letter].width = 16

        wb.save(output_path)
        print(f"[cohort] Excel saved → {output_path}")
    except ImportError:
        print("[cohort] openpyxl not installed — saving CSVs instead")
        for name, tbl in tables.items():
            tbl.to_csv(output_path.parent / f"{name}.csv")
