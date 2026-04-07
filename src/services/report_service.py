"""
services/report_service.py
===========================
Generates PDF and Excel reports from phase outputs.
"""
import json
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd

from core.config import get_settings
from services.segmentation_service import load_segment_results, load_global_insights
from services.drift_service import load_drift_report, load_early_warnings
from services.insight_service import load_latest_report
from utils.logger import log

settings = get_settings()


def generate_excel_report() -> Path:
    """Generate formatted Excel cohort report."""
    out_path = settings.REPORTS_DIR / f"cohort_report_{_ts()}.xlsx"
    settings.REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb    = openpyxl.Workbook()
        H_FILL = PatternFill("solid", fgColor="1D3557")
        H_FONT = Font(color="FFFFFF", bold=True)

        # Sheet 1: Segment summary
        ws = wb.active
        ws.title = "Segment Summary"
        headers = ["segment_id","churn_rate","prev_churn_rate","churn_delta",
                   "health_status","risk_level","revenue_at_risk","acceleration"]
        ws.append(headers)
        for c in ws[1]:
            c.fill = H_FILL; c.font = H_FONT

        for seg in load_segment_results():
            ws.append([seg.get(h) for h in headers])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        # Sheet 2: Drift features
        ws2 = wb.create_sheet("Drift Features")
        dh  = ["feature","psi","psi_status","drift_severity","trend","early_warning"]
        ws2.append(dh)
        for c in ws2[1]: c.fill = H_FILL; c.font = H_FONT
        dr  = load_drift_report()
        for f in dr.get("drift_report", []):
            ws2.append([f.get(h) for h in dh])
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 18

        # Sheet 3: Global insights
        ws3 = wb.create_sheet("Global Insights")
        ins  = load_global_insights()
        ws3.append(["Metric", "Value"])
        for c in ws3[1]: c.fill = H_FILL; c.font = H_FONT
        for k, v in ins.items():
            if not isinstance(v, (list, dict)):
                ws3.append([k, v])

        wb.save(out_path)
        log.info(f"[ReportService] Excel saved → {out_path}")
        return out_path

    except ImportError:
        # Fallback: plain CSV
        csv_path = settings.REPORTS_DIR / f"segment_report_{_ts()}.csv"
        pd.DataFrame(load_segment_results()).to_csv(csv_path, index=False)
        return csv_path


def generate_pdf_report() -> Path:
    """Generate PDF executive summary report."""
    out_path = settings.REPORTS_DIR / f"executive_report_{_ts()}.pdf"
    settings.REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle)

        doc    = SimpleDocTemplate(str(out_path), pagesize=A4)
        styles = getSampleStyleSheet()
        story  = []

        # Title
        story.append(Paragraph("Customer Health Forensics — Executive Report",
                                styles["Title"]))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                                styles["Normal"]))
        story.append(Spacer(1, 20))

        # Executive summary
        report = load_latest_report()
        if report:
            story.append(Paragraph("Executive Summary", styles["Heading1"]))
            story.append(Paragraph(report.get("executive_summary","N/A"),
                                   styles["Normal"]))
            story.append(Spacer(1, 12))

            # Business impact
            bi = report.get("business_impact", {})
            story.append(Paragraph("Business Impact", styles["Heading1"]))
            bi_data = [["Metric", "Value"],
                       ["Revenue at Risk",       f"${bi.get('total_annual_revenue_at_risk',0):,.0f}"],
                       ["Projected Loss",         f"${bi.get('projected_loss_if_no_action',0):,.0f}"],
                       ["Potential Recovery",     f"${bi.get('potential_recovery',0):,.0f}"],
                       ["Critical Customers",     str(bi.get('critical_customers_count',0))]]
            t = Table(bi_data, colWidths=[200, 200])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#1D3557")),
                ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
                ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
                ("GRID",       (0,0),(-1,-1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.white, colors.lightgrey]),
            ]))
            story.append(t)
            story.append(Spacer(1,12))

            # Top recommendations
            recs = report.get("recommendations",[])
            if recs:
                story.append(Paragraph("Top Recommendations & RL Actions", styles["Heading1"]))
                for r in recs:
                    story.append(Paragraph(
                        f"<b>{r.get('rank','?')}. Rank:</b> {r.get('description','—')} (Target: {r.get('target', '—')}) [Source: {r.get('source', '—')}]",
                        styles["Normal"]
                    ))
                    story.append(Spacer(1, 6))

            # Causal Logic & XAI
            causal = report.get("causal_analysis", {})
            if causal:
                story.append(Paragraph("Causality & XAI Insights (SHAP)", styles["Heading1"]))
                story.append(Paragraph(causal.get("narrative", "N/A"), styles["Normal"]))
                story.append(Spacer(1, 6))
                for driver in causal.get("top_xai_drivers", [])[:5]:
                     story.append(Paragraph(f"- <b>{driver.get('feature', '')}</b> (Importance: {driver.get('importance', 0):.2f}, Direction: {driver.get('direction', '')})", styles["Normal"]))
                story.append(Spacer(1, 12))

            # Segments Insight
            segs = report.get("segments", {})
            if segs:
                story.append(Paragraph("Segment Vulnerability", styles["Heading1"]))
                story.append(Paragraph(segs.get("narrative", "N/A"), styles["Normal"]))
                story.append(Spacer(1, 12))

            # Drift & Operational Readiness
            drift = report.get("drift_analysis", {})
            if drift:
                story.append(Paragraph("Data & Concept Drift Over Time", styles["Heading1"]))
                story.append(Paragraph(f"Overall Severity: <b>{drift.get('overall_severity', 'N/A')}</b>", styles["Normal"]))
                warnings_str = ", ".join(drift.get("early_warning_features", []))
                story.append(Paragraph(f"Early Warning Features: {warnings_str if warnings_str else 'None'}", styles["Normal"]))
                story.append(Spacer(1, 12))

        doc.build(story)
        log.info(f"[ReportService] PDF saved → {out_path}")
        return out_path

    except ImportError:
        # Fallback: plain text
        txt_path = settings.REPORTS_DIR / f"executive_report_{_ts()}.txt"
        report   = load_latest_report() or {}
        txt_path.write_text(report.get("executive_summary", "Report not available."))
        return txt_path


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")
